#Contactenates two excel files or sheets and create a new sheet
import pandas as pds
import numpy as np
import xlwings as xw
import openpyxl as op
# from openpyxl import Workbook
# from openpyxl.styles import Color, PatternFill, Font, Border
# from openpyxl.styles import colors
# from openpyxl.cell import Cell

file1 =('Vendors Consolidated.xlsx')
file2 =('Accounts.xlsx')

#Concatenate from two sheets / files
vendors_file = pds.read_excel(file1)
accounts_file = pds.read_excel(file2)
joinedData = pds.concat([vendors_file, accounts_file])
print("\n\t\tVendors Consolidated")
print(vendors_file)
print("\n\t\tAccounts")
print(accounts_file)

#Checking Vendors Consolidated for missing/error fields
for i in range(vendors_file.shape[0]): #iterate over rows
    for j in range(vendors_file.shape[1]): #iterate over columns
        #print("\nLocation: i = ",i,"j= ",j)
        if(vendors_file.iloc[i,j] == ""):
            print()
            #print("\n Empty value located at [",i,"] [",j,"]")
            # cell_obj = sheet_obj.cell(row = i+2, column = j+2)
            # cell_obj.font = op.styles.Font(color="00FF0000")
            # #cell_obj.font = op.styles.Font(name="Arial", size=16, color="00FF0000")
            # #sheet_obj.cell(row = i+1, column = j+1).font = Font(name="Arial", size=16, color="00FF0000")
            # print("Displaying object using openpyxl ",cell_obj.value)
            # print("\n Value check failed\n ")
            # discrepancies = discrepancies+1
    

#Concatenate tables
#joinedData = pds.concat([vendors_file, accounts_file])

#Merge Tables
joinedData = vendors_file.merge(accounts_file, how="inner", on='TRUCK NO')
#joinedData = vendors_file.merge(accounts_file, how="inner", on=['DATE','DC NO','TRUCK NO'])


print("\n\t\tOutput Table")
print (joinedData)
joinedData = joinedData.sort_index(axis=1)
joinedData.to_excel('Output Truck.xlsx')
output_vendors = joinedData[["DC NO","TRUCK NO","Received Qty_x","Accepted Qty_x"]]
# output_vendors = output_vendors.sort_index(axis=1)
print("\n\t\t Output Vendors")
print(output_vendors)
output_accounts = joinedData[["DC NO","TRUCK NO","Received Qty_y","Accepted Qty_y"]]
# output_accounts = output_accounts.sort_index(axis=1)
print("\n\t\t Output Accounts")
print(output_accounts)
output_accounts.to_excel('Output Accounts.xlsx')
output_vendors.to_excel('Output Vendors Consolidated.xlsx')

wb_obj = op.load_workbook("Output Accounts.xlsx")
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

 #Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
 
# Cell object is created by using
# sheet object's cell() method.
#cell_obj = sheet_obj.cell(row = 1, column = 1)
 
# Print value of cell object
# using the value attribute
#print(cell_obj.value)

#discrepancy_data = vendors_file.compare(accounts_file)
#print("\n\t\t Discrepancies table")
#fill_gen = op.styles.PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
print("Shape of Output Accounts ",output_accounts.shape)
discrepancies = 0
# for i in range(output_accounts.shape[0]): #iterate over rows
#     for j in range(output_accounts.shape[1]): #iterate over columns
#         print("\nLocation: i = ",i,"j= ",j)
#         if(output_accounts.iloc[i,j] != output_vendors.iloc[i,j]):
#             cell_obj = sheet_obj.cell(row = i+2, column = j+2)
#             cell_obj.font = op.styles.Font(color="00FF0000")
#             #cell_obj.font = op.styles.Font(name="Arial", size=16, color="00FF0000")
#             #sheet_obj.cell(row = i+1, column = j+1).font = Font(name="Arial", size=16, color="00FF0000")
#             print("Displaying object using openpyxl ",cell_obj.value)
#             print("\n Value check failed\n ")
#             discrepancies = discrepancies+1

#         else:
#             cell_obj = sheet_obj.cell(row = i+2, column = j+2)
#             print("Displaying object using openpyxl ",cell_obj.value)
#             print("\nPASSED\n")
#         # value = output_accounts.iloc[i, j] #get cell value
#         # print(value, end="\t")
#     print()

print("\nTOTAL DISCREPANCIES = ", discrepancies)
wb_obj.save("Discrepancies.xlsx")

# i=3
# j=2
# print("\n Cell Value for i = ",i," j = ",j," in Table Output Accounts")
# print(output_accounts.iloc[i,j])

# print("\n\t\t New Data Table")
# print(joinedData[['Country','Price']])
# c = np.where(vendors_file['Color'] == accounts_file['Color'])
# joinedData['Color Matched'] = np.where(vendors_file['Color'] == accounts_file['Color'],'True','False')
# print("\n\t")
# print(joinedData)

#To generate output file
#joinedData.to_excel('Output File.xlsx')
