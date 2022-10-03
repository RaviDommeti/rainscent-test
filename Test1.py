#Contactenates two excel files or sheets and create a new sheet
import pandas as pds
import numpy as np
import xlwings as xw
import openpyxl as op
# from openpyxl import Workbook
# from openpyxl.styles import Color, PatternFill, Font, Border
# from openpyxl.styles import colors
# from openpyxl.cell import Cell

file1 =('Vendors Consolidated.xlsx')
file2 =('Accounts.xlsx')

#Concatenate from two sheets / files
vendors_file = pds.read_excel(file1)
accounts_file = pds.read_excel(file2)
joinedData = pds.concat([vendors_file, accounts_file])
#print("\n\t\tVendors Consolidated")
#print(vendors_file)
# print("\n\t\tAccounts")
# print(accounts_file)
#print("\n Size of Vendors File: ",vendors_file.shape," with Rows: ",vendors_file.shape[0]," Columns: ",vendors_file.shape[1])
# Find rows and columns with null values
result_error = np.where(pds.isnull(vendors_file))
# result_error[0] has rows with null values while result_error[1] has columns with null value starting from 0
# print("\n Error in Rows: ",result_error[0],"\nError in Columns:  ",result_error[1])
rows_error = result_error[0] # Storing Error rows
cols_error = result_error[1] # Storing Error cols
len_rows_errors = len(result_error[0])
print("\nNumber of Rows with  errors:  ",len_rows_errors)
#Retrieve locations of error values
print("\nError Value Locations ")
for i in range(len_rows_errors):
    print("\n[",rows_error[i],"] [",cols_error[i],"]")
# print("\n[",result_error[0][0],"] [",result_error[1][0],"]\n")
# print("\n[",rows_error[0],"] [",cols_error[0],"]")
#Storing column names of Vendors consolidated
cols_list = list(vendors_file.columns)
print("\n Vendors Consolidated Columns: ",cols_list,"\n")
#Finding number of columsn in Vendors consolidated starts from 1
total_cols_vendors = len(cols_list) 
print("\n Number of Columns in Vendors Consolidated",total_cols_vendors,"\n")
# Duplicating vendors file
dup_vendors_file = vendors_file
print("\n Number of Columns in Vendors Duplicated",len(list(dup_vendors_file.columns)),"\n")
dup_vendors_file['Errors'] = ""
print("\n Number of Columns in New Vendors Duplicated",len(list(dup_vendors_file.columns)),"\n")
# print("\n\t\tDuplicate Vendors Consolidated")
# print(dup_vendors_file)
error_string = []
for i in  range(3):
    error_string.append("i= "+str(i))
print("\nError String: ",error_string)

# testprint = np.where(pds.isnull(vendors_file))
# print("\n ",testprint)
# https://stackoverflow.com/questions/27159189/find-empty-or-nan-entry-in-pandas-dataframe
#Checking Vendors Consolidated for missing/error fields
# for i in range(vendors_file.shape[0]): #iterate over rows from i =0
#     for j in range(vendors_file.shape[1]): #iterate over columns j = 0....[0][0] points to Row 1 Column 1(excludes row id and column name)
#         #print("\nValue at [",i,"] [",j,"] is ",vendors_file.iloc[i,j])
#         if(vendors_file.iloc[i,j] == "NaN" or vendors_file.iloc[i,j] == "" or vendors_file.iloc[i,j] == " "):
#             print("\nValue at [",i,"] [",j,"] is ",vendors_file.iloc[i,j])
#         #print("\nLocation: i = ",i,"j= ",j)
#         # if(vendors_file.iloc[i,j] == "Nan" or vendors_file.iloc[i,j] == "NaT"):
#         #     print("\nLocation: i = ",i,"j= ",j)
#         #     print("Empty value")
#         #     print("\n Empty value located at [",i,"] [",j,"]")
#         #     cell_obj = sheet_obj.cell(row = i+2, column = j+2)
#         #     cell_obj.font = op.styles.Font(color="00FF0000")
#         #     #cell_obj.font = op.styles.Font(name="Arial", size=16, color="00FF0000")
#         #     #sheet_obj.cell(row = i+1, column = j+1).font = Font(name="Arial", size=16, color="00FF0000")
#         #     print("Displaying object using openpyxl ",cell_obj.value)
#         #     print("\n Value check failed\n ")
#         #     discrepancies = discrepancies+1
    

#Concatenate tables
#joinedData = pds.concat([vendors_file, accounts_file])

#Merge Tables
#joinedData = vendors_file.merge(accounts_file, how="inner", on='TRUCK NO')
#joinedData = vendors_file.merge(accounts_file, how="inner", on=['DATE','DC NO','TRUCK NO'])


# print("\n\t\tOutput Table")
# print (joinedData)
# joinedData = joinedData.sort_index(axis=1)
# joinedData.to_excel('Output Truck.xlsx')
# output_vendors = joinedData[["DC NO","TRUCK NO","Received Qty_x","Accepted Qty_x"]]
# print("\n\t\t Output Vendors")
# print(output_vendors)
# output_accounts = joinedData[["DC NO","TRUCK NO","Received Qty_y","Accepted Qty_y"]]
# print("\n\t\t Output Accounts")
# print(output_accounts)
# output_accounts.to_excel('Output Accounts.xlsx')
# output_vendors.to_excel('Output Vendors Consolidated.xlsx')



